import pandas as pd
import logging
from datetime import datetime, timedelta
from typing import List, Dict
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from models import DatabaseManager, TradeRecord
from config import Config

logger = logging.getLogger(__name__)

class MacadamiaTradeExcelReporter:
    def __init__(self):
        self.config = Config()
        self.db = DatabaseManager(self.config.DATABASE_URL)
    
    def generate_daily_excel_report(self, date_str: str = None) -> str:
        """일일 엑셀 보고서 생성"""
        if date_str is None:
            date_str = datetime.now().strftime('%Y%m%d')
        
        try:
            # 최근 데이터 가져오기
            records = self.db.get_latest_records(1)
            
            if not records:
                logger.warning("보고서 생성할 데이터가 없습니다.")
                return None
            
            # 엑셀 파일 생성
            filename = f"reports/macadamia_trade_report_{date_str}.xlsx"
            os.makedirs("reports", exist_ok=True)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # 1. 원본 데이터 시트
                self._create_raw_data_sheet(records, writer)
                
                # 2. 국가별 요약 시트
                self._create_country_summary_sheet(records, writer)
                
                # 3. 제품별 요약 시트
                self._create_product_summary_sheet(records, writer)
                
                # 4. 월별 트렌드 시트
                self._create_monthly_trend_sheet(writer)
                
                # 5. 데이터 입력 템플릿 시트
                self._create_input_template_sheet(writer)
            
            # 엑셀 파일 스타일링
            self._apply_excel_styling(filename)
            
            logger.info(f"엑셀 보고서 생성 완료: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"엑셀 보고서 생성 오류: {e}")
            return None
    
    def _create_raw_data_sheet(self, records: List[TradeRecord], writer):
        """원본 데이터 시트 생성"""
        data = []
        for record in records:
            data.append({
                'ID': record.id,
                '날짜': record.date,
                '수출국': record.country_origin,
                '수입국': record.country_destination,
                '수출업체': record.company_exporter or '',
                '수입업체': record.company_importer or '',
                '제품코드': record.product_code,
                '제품명': record.product_description,
                '수량': record.quantity,
                '단위': record.unit,
                '금액(USD)': record.value_usd,
                '거래유형': record.trade_type,
                '등록일시': record.created_at
            })
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='원본데이터', index=False)
    
    def _create_country_summary_sheet(self, records: List[TradeRecord], writer):
        """국가별 요약 시트 생성"""
        country_data = {}
        
        for record in records:
            country = record.country_origin
            if country not in country_data:
                country_data[country] = {
                    '수출국': country,
                    '거래건수': 0,
                    '총금액(USD)': 0,
                    '평균금액(USD)': 0,
                    '총수량': 0,
                    '주요제품': set()
                }
            
            country_data[country]['거래건수'] += 1
            country_data[country]['총금액(USD)'] += record.value_usd or 0
            country_data[country]['총수량'] += record.quantity or 0
            country_data[country]['주요제품'].add(record.product_description or '')
        
        # 평균 계산 및 주요제품 문자열 변환
        summary_data = []
        for country, data in country_data.items():
            data['평균금액(USD)'] = data['총금액(USD)'] / data['거래건수'] if data['거래건수'] > 0 else 0
            data['주요제품'] = ', '.join(list(data['주요제품'])[:3])  # 상위 3개만
            summary_data.append(data)
        
        df = pd.DataFrame(summary_data)
        df = df.sort_values('총금액(USD)', ascending=False)
        df.to_excel(writer, sheet_name='국가별요약', index=False)
    
    def _create_product_summary_sheet(self, records: List[TradeRecord], writer):
        """제품별 요약 시트 생성"""
        product_data = {}
        
        for record in records:
            product = record.product_description or '미분류'
            if product not in product_data:
                product_data[product] = {
                    '제품명': product,
                    '제품코드': record.product_code,
                    '거래건수': 0,
                    '총금액(USD)': 0,
                    '평균단가(USD/kg)': 0,
                    '총수량(kg)': 0,
                    '주요수출국': set()
                }
            
            product_data[product]['거래건수'] += 1
            product_data[product]['총금액(USD)'] += record.value_usd or 0
            product_data[product]['총수량(kg)'] += record.quantity or 0
            product_data[product]['주요수출국'].add(record.country_origin or '')
        
        # 평균 단가 계산
        summary_data = []
        for product, data in product_data.items():
            if data['총수량(kg)'] > 0:
                data['평균단가(USD/kg)'] = data['총금액(USD)'] / data['총수량(kg)']
            data['주요수출국'] = ', '.join(list(data['주요수출국'])[:3])
            summary_data.append(data)
        
        df = pd.DataFrame(summary_data)
        df = df.sort_values('총금액(USD)', ascending=False)
        df.to_excel(writer, sheet_name='제품별요약', index=False)
    
    def _create_monthly_trend_sheet(self, writer):
        """월별 트렌드 시트 생성"""
        try:
            # 최근 12개월 데이터 가져오기
            records = self.db.get_latest_records(365)
            
            monthly_data = {}
            for record in records:
                month_key = record.date.strftime('%Y-%m') if record.date else 'Unknown'
                if month_key not in monthly_data:
                    monthly_data[month_key] = {
                        '년월': month_key,
                        '거래건수': 0,
                        '총금액(USD)': 0,
                        '총수량(kg)': 0,
                        '평균단가(USD/kg)': 0
                    }
                
                monthly_data[month_key]['거래건수'] += 1
                monthly_data[month_key]['총금액(USD)'] += record.value_usd or 0
                monthly_data[month_key]['총수량(kg)'] += record.quantity or 0
            
            # 평균 단가 계산
            trend_data = []
            for month, data in monthly_data.items():
                if data['총수량(kg)'] > 0:
                    data['평균단가(USD/kg)'] = data['총금액(USD)'] / data['총수량(kg)']
                trend_data.append(data)
            
            df = pd.DataFrame(trend_data)
            df = df.sort_values('년월')
            df.to_excel(writer, sheet_name='월별트렌드', index=False)
            
        except Exception as e:
            logger.error(f"월별 트렌드 시트 생성 오류: {e}")
    
    def _create_input_template_sheet(self, writer):
        """데이터 입력 템플릿 시트 생성"""
        template_data = [
            {
                '날짜': '2025-01-10',
                '수출국': '호주',
                '수입국': '한국',
                '수출업체': '예: Australian Macadamia Co.',
                '수입업체': '예: Korea Nuts Import Ltd.',
                '제품코드': '080250',
                '제품명': '예: Raw Macadamia Nuts',
                '수량': 1000,
                '단위': 'kg',
                '금액(USD)': 15000,
                '거래유형': 'export',
                '비고': '신규 데이터 입력 시 이 행을 복사하여 사용'
            }
        ]
        
        df = pd.DataFrame(template_data)
        df.to_excel(writer, sheet_name='데이터입력템플릿', index=False)
    
    def _apply_excel_styling(self, filename: str):
        """엑셀 파일에 스타일 적용"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(filename)
            
            # 헤더 스타일 정의
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 헤더 행 스타일 적용
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 열 너비 자동 조정
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            
        except Exception as e:
            logger.error(f"엑셀 스타일링 오류: {e}")
    
    def create_comparison_report(self, days: int = 7) -> str:
        """비교 분석 엑셀 보고서 생성"""
        try:
            filename = f"reports/macadamia_comparison_{datetime.now().strftime('%Y%m%d')}.xlsx"
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # 기간별 비교 데이터
                current_records = self.db.get_latest_records(days)
                previous_records = self.db.get_records_by_date_range(
                    datetime.now() - timedelta(days=days*2),
                    datetime.now() - timedelta(days=days)
                )
                
                # 현재 기간 데이터
                self._create_period_summary(current_records, writer, f'최근{days}일')
                
                # 이전 기간 데이터
                self._create_period_summary(previous_records, writer, f'이전{days}일')
                
                # 비교 분석
                self._create_comparison_analysis(current_records, previous_records, writer)
            
            self._apply_excel_styling(filename)
            return filename
            
        except Exception as e:
            logger.error(f"비교 보고서 생성 오류: {e}")
            return None
    
    def _create_period_summary(self, records: List[TradeRecord], writer, sheet_name: str):
        """기간별 요약 시트 생성"""
        if not records:
            return
        
        summary_data = []
        country_stats = {}
        
        for record in records:
            country = record.country_origin
            if country not in country_stats:
                country_stats[country] = {
                    '수출국': country,
                    '거래건수': 0,
                    '총금액(USD)': 0,
                    '총수량(kg)': 0
                }
            
            country_stats[country]['거래건수'] += 1
            country_stats[country]['총금액(USD)'] += record.value_usd or 0
            country_stats[country]['총수량(kg)'] += record.quantity or 0
        
        for data in country_stats.values():
            summary_data.append(data)
        
        df = pd.DataFrame(summary_data)
        df = df.sort_values('총금액(USD)', ascending=False)
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    def _create_comparison_analysis(self, current_records: List[TradeRecord], previous_records: List[TradeRecord], writer):
        """비교 분석 시트 생성"""
        current_total = sum(record.value_usd or 0 for record in current_records)
        previous_total = sum(record.value_usd or 0 for record in previous_records)
        
        growth_rate = ((current_total - previous_total) / previous_total * 100) if previous_total > 0 else 0
        
        analysis_data = [
            {
                '항목': '총 거래액(USD)',
                '현재 기간': current_total,
                '이전 기간': previous_total,
                '증감액': current_total - previous_total,
                '증감률(%)': round(growth_rate, 2)
            },
            {
                '항목': '총 거래 건수',
                '현재 기간': len(current_records),
                '이전 기간': len(previous_records),
                '증감액': len(current_records) - len(previous_records),
                '증감률(%)': round(((len(current_records) - len(previous_records)) / len(previous_records) * 100), 2) if len(previous_records) > 0 else 0
            }
        ]
        
        df = pd.DataFrame(analysis_data)
        df.to_excel(writer, sheet_name='비교분석', index=False)
    
    def create_enhanced_report(self, detailed_data: List[Dict], start_date, end_date, include_detailed_analysis=True) -> str:
        """상세 시뮬레이션 데이터를 활용한 향상된 리포트 생성"""
        
        if not detailed_data:
            return "데이터가 없어 리포트를 생성할 수 없습니다."
        
        # 리포트 내용 생성
        report_lines = []
        
        # 헤더
        report_lines.extend([
            "# 마카다미아 무역 데이터 상세 분석 리포트",
            f"**생성일시**: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M:%S')}",
            f"**분석기간**: {start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}",
            f"**총 거래건수**: {len(detailed_data)}건",
            "",
            "---",
            ""
        ])
        
        # 1. 수입업체 분석
        report_lines.extend(self._analyze_importers(detailed_data))
        
        # 2. 수출업체 및 원산지 분석
        report_lines.extend(self._analyze_exporters(detailed_data))
        
        # 3. 품목 및 품질 분석
        report_lines.extend(self._analyze_products(detailed_data))
        
        # 4. 가격 및 시장 분석
        report_lines.extend(self._analyze_pricing(detailed_data))
        
        # 5. 물류 및 통관 분석
        report_lines.extend(self._analyze_logistics(detailed_data))
        
        # 6. 규정 및 인증 분석
        report_lines.extend(self._analyze_regulations(detailed_data))
        
        if include_detailed_analysis:
            # 7. 상세 거래 내역
            report_lines.extend(self._create_detailed_transactions(detailed_data))
        
        # 8. 시사점 및 권고사항
        report_lines.extend(self._create_recommendations(detailed_data))
        
        return '\n'.join(report_lines)
    
    def _analyze_importers(self, data: List[Dict]) -> List[str]:
        """수입업체 분석"""
        lines = [
            "## 1. 수입업체 현황 분석",
            ""
        ]
        
        # 수입업체별 통계
        importer_stats = {}
        for record in data:
            company = record.get('company_importer', '알 수 없음')
            if company not in importer_stats:
                importer_stats[company] = {
                    'count': 0,
                    'total_value': 0,
                    'total_quantity': 0,
                    'business_type': record.get('importer_business_type', ''),
                    'specialization': record.get('importer_specialization', ''),
                    'established': record.get('importer_established', ''),
                    'employees': record.get('importer_employees', ''),
                    'address': record.get('importer_address', '')
                }
            
            importer_stats[company]['count'] += 1
            importer_stats[company]['total_value'] += float(record.get('value_usd', 0))
            importer_stats[company]['total_quantity'] += float(record.get('quantity', 0))
        
        lines.append("### 주요 수입업체 현황")
        lines.append("| 순위 | 업체명 | 거래건수 | 총 수입액(USD) | 총 수입량(kg) | 업종 |")
        lines.append("|------|--------|----------|----------------|----------------|------|")
        
        # 수입액 기준 정렬
        sorted_importers = sorted(importer_stats.items(), key=lambda x: x[1]['total_value'], reverse=True)
        
        for i, (company, stats) in enumerate(sorted_importers[:10], 1):
            lines.append(f"| {i} | {company} | {stats['count']}건 | ${stats['total_value']:,.0f} | {stats['total_quantity']:,.0f} | {stats['business_type']} |")
        
        lines.extend([
            "",
            "### 수입업체 특성 분석",
            ""
        ])
        
        # 업종별 분류
        business_types = {}
        for record in data:
            btype = record.get('importer_business_type', '기타')
            business_types[btype] = business_types.get(btype, 0) + 1
        
        lines.append("**업종별 분포:**")
        for btype, count in sorted(business_types.items(), key=lambda x: x[1], reverse=True):
            percentage = (count / len(data)) * 100
            lines.append(f"- {btype}: {count}건 ({percentage:.1f}%)")
        
        lines.extend(["", "---", ""])
        return lines
    
    def _analyze_exporters(self, data: List[Dict]) -> List[str]:
        """수출업체 및 원산지 분석"""
        lines = [
            "## 2. 수출업체 및 원산지 분석",
            ""
        ]
        
        # 원산지별 통계
        country_stats = {}
        exporter_stats = {}
        
        for record in data:
            country = record.get('country_origin', '알 수 없음')
            exporter = record.get('company_exporter', '알 수 없음')
            
            # 원산지 통계
            if country not in country_stats:
                country_stats[country] = {'count': 0, 'value': 0, 'quantity': 0}
            country_stats[country]['count'] += 1
            country_stats[country]['value'] += float(record.get('value_usd', 0))
            country_stats[country]['quantity'] += float(record.get('quantity', 0))
            
            # 수출업체 통계
            if exporter not in exporter_stats:
                exporter_stats[exporter] = {
                    'count': 0, 'value': 0, 'quantity': 0,
                    'country': country,
                    'certification': record.get('exporter_certification', ''),
                    'capacity': record.get('exporter_capacity', ''),
                    'experience': record.get('exporter_experience', '')
                }
            exporter_stats[exporter]['count'] += 1
            exporter_stats[exporter]['value'] += float(record.get('value_usd', 0))
            exporter_stats[exporter]['quantity'] += float(record.get('quantity', 0))
        
        # 원산지별 현황
        lines.append("### 원산지별 수입 현황")
        lines.append("| 순위 | 원산지 | 거래건수 | 수입액(USD) | 수입량(kg) | 점유율 |")
        lines.append("|------|--------|----------|-------------|------------|--------|")
        
        total_value = sum(stats['value'] for stats in country_stats.values())
        sorted_countries = sorted(country_stats.items(), key=lambda x: x[1]['value'], reverse=True)
        
        for i, (country, stats) in enumerate(sorted_countries, 1):
            share = (stats['value'] / total_value) * 100 if total_value > 0 else 0
            lines.append(f"| {i} | {country} | {stats['count']}건 | ${stats['value']:,.0f} | {stats['quantity']:,.0f} | {share:.1f}% |")
        
        # 주요 수출업체
        lines.extend([
            "",
            "### 주요 수출업체 현황",
            "| 순위 | 업체명 | 원산지 | 거래건수 | 수입액(USD) | 주요 인증 |",
            "|------|--------|--------|----------|-------------|-----------|"
        ])
        
        sorted_exporters = sorted(exporter_stats.items(), key=lambda x: x[1]['value'], reverse=True)
        for i, (exporter, stats) in enumerate(sorted_exporters[:8], 1):
            cert = stats['certification'][:30] + "..." if len(stats['certification']) > 30 else stats['certification']
            lines.append(f"| {i} | {exporter} | {stats['country']} | {stats['count']}건 | ${stats['value']:,.0f} | {cert} |")
        
        lines.extend(["", "---", ""])
        return lines
    
    def _analyze_products(self, data: List[Dict]) -> List[str]:
        """품목 및 품질 분석"""
        lines = [
            "## 3. 품목 및 품질 분석",
            ""
        ]
        
        # HS코드별 통계
        hs_stats = {}
        quality_stats = {}
        
        for record in data:
            hs_code = record.get('product_code', '알 수 없음')
            quality = record.get('quality_grade', '알 수 없음')
            
            # HS코드 통계
            if hs_code not in hs_stats:
                hs_stats[hs_code] = {
                    'count': 0, 'value': 0, 'quantity': 0,
                    'description': record.get('product_description', ''),
                    'detailed_desc': record.get('product_detailed_description', '')
                }
            hs_stats[hs_code]['count'] += 1
            hs_stats[hs_code]['value'] += float(record.get('value_usd', 0))
            hs_stats[hs_code]['quantity'] += float(record.get('quantity', 0))
            
            # 품질등급 통계
            quality_stats[quality] = quality_stats.get(quality, 0) + 1
        
        # HS코드별 현황
        lines.append("### HS코드별 수입 현황")
        lines.append("| HS코드 | 품목명 | 거래건수 | 수입액(USD) | 수입량(kg) | 평균단가(USD/kg) |")
        lines.append("|--------|--------|----------|-------------|------------|------------------|")
        
        for hs_code, stats in sorted(hs_stats.items(), key=lambda x: x[1]['value'], reverse=True):
            avg_price = stats['value'] / stats['quantity'] if stats['quantity'] > 0 else 0
            lines.append(f"| {hs_code} | {stats['description']} | {stats['count']}건 | ${stats['value']:,.0f} | {stats['quantity']:,.0f} | ${avg_price:.2f} |")
        
        # 품질등급 분포
        lines.extend([
            "",
            "### 품질등급별 분포",
            ""
        ])
        
        total_records = len(data)
        for quality, count in sorted(quality_stats.items(), key=lambda x: x[1], reverse=True):
            percentage = (count / total_records) * 100
            lines.append(f"- **{quality}**: {count}건 ({percentage:.1f}%)")
        
        lines.extend(["", "---", ""])
        return lines
    
    def _analyze_pricing(self, data: List[Dict]) -> List[str]:
        """가격 및 시장 분석"""
        lines = [
            "## 4. 가격 및 시장 분석",
            ""
        ]
        
        # 가격 통계 계산
        prices = []
        values = []
        quantities = []
        
        for record in data:
            price = float(record.get('unit_price_usd', 0))
            value = float(record.get('value_usd', 0))
            quantity = float(record.get('quantity', 0))
            
            if price > 0:
                prices.append(price)
            if value > 0:
                values.append(value)
            if quantity > 0:
                quantities.append(quantity)
        
        if prices:
            lines.extend([
                "### 가격 동향 분석",
                f"- **평균 단가**: ${sum(prices)/len(prices):.2f}/kg",
                f"- **최고 단가**: ${max(prices):.2f}/kg",
                f"- **최저 단가**: ${min(prices):.2f}/kg",
                f"- **가격 편차**: ${(max(prices) - min(prices)):.2f}/kg",
                ""
            ])
        
        # 거래규모 분석
        if values:
            lines.extend([
                "### 거래규모 분석",
                f"- **총 거래액**: ${sum(values):,.0f}",
                f"- **평균 거래액**: ${sum(values)/len(values):,.0f}",
                f"- **최대 거래액**: ${max(values):,.0f}",
                f"- **최소 거래액**: ${min(values):,.0f}",
                ""
            ])
        
        if quantities:
            lines.extend([
                "### 수입량 분석",
                f"- **총 수입량**: {sum(quantities):,.0f} kg",
                f"- **평균 수입량**: {sum(quantities)/len(quantities):,.0f} kg",
                f"- **최대 수입량**: {max(quantities):,.0f} kg",
                f"- **최소 수입량**: {min(quantities):,.0f} kg",
                ""
            ])
        
        # 관세 및 세금 분석
        tariff_analysis = self._analyze_tariffs(data)
        lines.extend(tariff_analysis)
        
        lines.extend(["---", ""])
        return lines
    
    def _analyze_tariffs(self, data: List[Dict]) -> List[str]:
        """관세 및 세금 분석"""
        lines = [
            "### 관세 및 세금 분석",
            ""
        ]
        
        tariff_rates = {}
        total_tariff = 0
        total_vat = 0
        total_value = 0
        
        for record in data:
            rate = record.get('tariff_rate_applied', '0%')
            tariff_amount = float(record.get('tariff_amount', 0))
            vat_amount = float(record.get('vat_amount', 0))
            value = float(record.get('value_usd', 0))
            
            tariff_rates[rate] = tariff_rates.get(rate, 0) + 1
            total_tariff += tariff_amount
            total_vat += vat_amount
            total_value += value
        
        lines.append("**적용 관세율 분포:**")
        for rate, count in sorted(tariff_rates.items()):
            lines.append(f"- {rate}: {count}건")
        
        lines.extend([
            "",
            "**세금 부담 현황:**",
            f"- 총 관세액: ${total_tariff:,.0f}",
            f"- 총 부가세: ${total_vat:,.0f}",
            f"- 세금 부담률: {((total_tariff + total_vat) / total_value * 100):.1f}%" if total_value > 0 else "- 세금 부담률: 0%",
            ""
        ])
        
        return lines
    
    def _analyze_logistics(self, data: List[Dict]) -> List[str]:
        """물류 및 통관 분석"""
        lines = [
            "## 5. 물류 및 통관 분석",
            ""
        ]
        
        # 항구별 통계
        port_stats = {}
        container_stats = {}
        shipping_stats = {}
        
        for record in data:
            loading_port = record.get('port_of_loading', '알 수 없음')
            discharge_port = record.get('port_of_discharge', '알 수 없음')
            container_type = record.get('container_type', '알 수 없음')
            shipping_line = record.get('shipping_line', '알 수 없음')
            
            # 항구 통계
            port_key = f"{loading_port} → {discharge_port}"
            port_stats[port_key] = port_stats.get(port_key, 0) + 1
            
            # 컨테이너 통계
            container_stats[container_type] = container_stats.get(container_type, 0) + 1
            
            # 선사 통계
            shipping_stats[shipping_line] = shipping_stats.get(shipping_line, 0) + 1
        
        # 주요 운송경로
        lines.extend([
            "### 주요 운송경로",
            "| 순위 | 운송경로 | 이용건수 |",
            "|------|----------|----------|"
        ])
        
        sorted_ports = sorted(port_stats.items(), key=lambda x: x[1], reverse=True)
        for i, (route, count) in enumerate(sorted_ports[:8], 1):
            lines.append(f"| {i} | {route} | {count}건 |")
        
        # 컨테이너 유형별 분포
        lines.extend([
            "",
            "### 컨테이너 유형별 분포",
            ""
        ])
        
        total_containers = sum(container_stats.values())
        for container_type, count in sorted(container_stats.items(), key=lambda x: x[1], reverse=True):
            percentage = (count / total_containers) * 100
            lines.append(f"- **{container_type}**: {count}건 ({percentage:.1f}%)")
        
        # 선사별 분포
        lines.extend([
            "",
            "### 주요 선사별 이용 현황",
            ""
        ])
        
        for shipping_line, count in sorted(shipping_stats.items(), key=lambda x: x[1], reverse=True)[:5]:
            percentage = (count / len(data)) * 100
            lines.append(f"- **{shipping_line}**: {count}건 ({percentage:.1f}%)")
        
        lines.extend(["", "---", ""])
        return lines
    
    def _analyze_regulations(self, data: List[Dict]) -> List[str]:
        """규정 및 인증 분석"""
        lines = [
            "## 6. 규정 및 인증 분석",
            ""
        ]
        
        # 검사기관 분석
        inspection_counts = {}
        cert_analysis = {
            'organic': 0,
            'fair_trade': 0,
            'haccp': 0
        }
        
        for record in data:
            # 검사기관
            agencies = record.get('inspection_agencies', [])
            if isinstance(agencies, list):
                for agency in agencies:
                    inspection_counts[agency] = inspection_counts.get(agency, 0) + 1
            
            # 인증 분석
            if record.get('organic_certified'):
                cert_analysis['organic'] += 1
            if record.get('fair_trade_certified'):
                cert_analysis['fair_trade'] += 1
            if record.get('haccp_certified'):
                cert_analysis['haccp'] += 1
        
        # 주요 검사기관
        lines.extend([
            "### 주요 검사기관별 검사 현황",
            ""
        ])
        
        for agency, count in sorted(inspection_counts.items(), key=lambda x: x[1], reverse=True):
            percentage = (count / len(data)) * 100
            lines.append(f"- **{agency}**: {count}건 ({percentage:.1f}%)")
        
        # 인증 현황
        lines.extend([
            "",
            "### 인증 현황 분석",
            ""
        ])
        
        total_records = len(data)
        cert_names = {
            'organic': '유기농 인증',
            'fair_trade': '공정무역 인증',
            'haccp': 'HACCP 인증'
        }
        
        for cert_type, count in cert_analysis.items():
            percentage = (count / total_records) * 100
            lines.append(f"- **{cert_names[cert_type]}**: {count}건 ({percentage:.1f}%)")
        
        lines.extend(["", "---", ""])
        return lines
    
    def _create_detailed_transactions(self, data: List[Dict]) -> List[str]:
        """상세 거래 내역"""
        lines = [
            "## 7. 상세 거래 내역",
            "",
            "### 최근 주요 거래 현황",
            "| 신고번호 | 날짜 | 수입업체 | 수출업체 | 원산지 | 품목 | 수량(kg) | 금액(USD) |",
            "|----------|------|----------|----------|--------|------|----------|-----------|"
        ]
        
        # 금액 기준으로 정렬하여 상위 15건 표시
        sorted_data = sorted(data, key=lambda x: float(x.get('value_usd', 0)), reverse=True)
        
        for record in sorted_data[:15]:
            decl_num = record.get('declaration_number', '')
            date = record.get('date', '')
            importer = record.get('company_importer', '')[:20] + "..." if len(record.get('company_importer', '')) > 20 else record.get('company_importer', '')
            exporter = record.get('company_exporter', '')[:20] + "..." if len(record.get('company_exporter', '')) > 20 else record.get('company_exporter', '')
            origin = record.get('country_origin', '')
            product = record.get('product_code', '')
            quantity = f"{float(record.get('quantity', 0)):,.0f}"
            value = f"${float(record.get('value_usd', 0)):,.0f}"
            
            lines.append(f"| {decl_num} | {date} | {importer} | {exporter} | {origin} | {product} | {quantity} | {value} |")
        
        lines.extend(["", "---", ""])
        return lines
    
    def _create_recommendations(self, data: List[Dict]) -> List[str]:
        """시사점 및 권고사항"""
        lines = [
            "## 8. 시사점 및 권고사항",
            ""
        ]
        
        # 데이터 기반 인사이트 생성
        total_value = sum(float(record.get('value_usd', 0)) for record in data)
        avg_price = sum(float(record.get('unit_price_usd', 0)) for record in data if record.get('unit_price_usd', 0) > 0) / len([r for r in data if r.get('unit_price_usd', 0) > 0])
        
        # 원산지 다양성
        origin_countries = set(record.get('country_origin') for record in data)
        
        # FTA 활용률
        fta_records = len([r for r in data if r.get('tariff_rate_applied', '8%') == '0%'])
        fta_rate = (fta_records / len(data)) * 100
        
        lines.extend([
            "### 주요 시사점",
            "",
            f"1. **시장 규모**: 분석 기간 중 총 거래액은 ${total_value:,.0f}로 활발한 거래가 이루어지고 있음",
            f"2. **가격 수준**: 평균 단가 ${avg_price:.2f}/kg로 프리미엄 시장 형성",
            f"3. **공급 다양성**: {len(origin_countries)}개국에서 수입하여 공급원 다변화 실현",
            f"4. **FTA 활용**: 전체 거래의 {fta_rate:.1f}%가 FTA 특혜관세 적용",
            "",
            "### 권고사항",
            "",
            "#### 수입업체 대상",
            "- 품질등급별 차별화된 마케팅 전략 수립 필요",
            "- FTA 특혜관세 활용 확대로 비용 절감 효과 극대화",
            "- 유기농, HACCP 등 인증 제품 비중 확대 검토",
            "",
            "#### 정책 당국 대상",
            "- 마카다미아 수입 증가 추세에 따른 검역체계 효율화 필요",
            "- 원산지 다변화에 따른 품질관리 기준 표준화 검토",
            "- 프리미엄 견과류 시장 육성을 위한 지원책 마련",
            "",
            "---",
            "",
            f"**리포트 생성 완료** | 분석 데이터: {len(data)}건 | 생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            ""
        ])
        
        return lines
