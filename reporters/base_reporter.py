"""
기본 보고서 생성 클래스 및 공통 유틸리티
"""
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    
import logging
from datetime import datetime, timedelta
from typing import List, Dict
import os
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
from models import DatabaseManager, TradeRecord
from config import Config

logger = logging.getLogger(__name__)

class BaseReporter:
    """기본 보고서 생성 클래스"""
    
    def __init__(self):
        self.config = Config()
        self.db = DatabaseManager(self.config.DATABASE_URL)
        
        # Excel 스타일 정의 (openpyxl이 있는 경우만)
        if OPENPYXL_AVAILABLE:
            from openpyxl.styles import Font, PatternFill, Alignment
            self.header_font = Font(bold=True, color="FFFFFF")
            self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            self.center_alignment = Alignment(horizontal="center", vertical="center")
        else:
            self.header_font = None
            self.header_fill = None
            self.center_alignment = None
        
    def _ensure_reports_directory(self):
        """보고서 디렉토리 생성"""
        os.makedirs("reports", exist_ok=True)
        
    def _apply_header_style(self, worksheet, max_col: int):
        """헤더 스타일 적용"""
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            
    def _format_currency(self, value: float) -> str:
        """통화 포맷팅"""
        if value is None or value == 0:
            return "$0"
        return f"${value:,.2f}"
        
    def _format_quantity(self, value: float) -> str:
        """수량 포맷팅"""
        if value is None or value == 0:
            return "0 kg"
        return f"{value:,.0f} kg"
        
    def _get_date_range_records(self, start_date: datetime, end_date: datetime) -> List[TradeRecord]:
        """날짜 범위 내 레코드 조회"""
        try:
            # DatabaseManager에 날짜 범위 조회 메서드가 있다고 가정
            return self.db.get_records_by_date_range(start_date, end_date)
        except Exception as e:
            logger.error(f"날짜 범위 조회 오류: {e}")
            # 폴백: 최근 레코드 조회
            return self.db.get_latest_records(100)
            
    def _create_dataframe_from_records(self, records: List[TradeRecord]):
        """TradeRecord 리스트를 DataFrame으로 변환"""
        if not PANDAS_AVAILABLE:
            raise ImportError("pandas is not available")
            
        data = []
        for record in records:
            data.append({
                'Date': record.date,
                'Product Code': record.product_code,
                'Product Description': record.product_description,
                'Origin Country': record.country_origin,
                'Destination Country': record.country_destination,
                'Trade Value (USD)': record.trade_value,
                'Quantity (kg)': record.quantity,
                'Trade Type': record.trade_type,
                'Period': record.period,
                'Year': record.year,
                'Source': record.source
            })
        return pd.DataFrame(data)
        
    def close(self):
        """리소스 정리"""
        if self.db:
            self.db.close()
