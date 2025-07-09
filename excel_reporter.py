import pandas as pd
import logging
from datetime import datetime, timedelta
from typing import List, Dict
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from models import DatabaseManager, TradeRecord
from config import Config

logger = logging.getLogger(__name__)

class MacadamiaTradeExcelReporter:
    def __init__(self):
        self.config = Config()
        self.db = DatabaseManager(self.config.DATABASE_URL)
    
    def generate_daily_excel_report(self, date_str: str = None) -> str:
        """일일 엑셀 보고서 생성"""
        if date_str is None:
            date_str = datetime.now().strftime('%Y%m%d')
        
        try:
            # 최근 데이터 가져오기
            records = self.db.get_latest_records(1)
            
            if not records:
                logger.warning("보고서 생성할 데이터가 없습니다.")
                return None
            
            # 엑셀 파일 생성
            filename = f"reports/macadamia_trade_report_{date_str}.xlsx"
            os.makedirs("reports", exist_ok=True)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # 1. 원본 데이터 시트
                self._create_raw_data_sheet(records, writer)
                
                # 2. 국가별 요약 시트
                self._create_country_summary_sheet(records, writer)
                
                # 3. 제품별 요약 시트
                self._create_product_summary_sheet(records, writer)
                
                # 4. 월별 트렌드 시트
                self._create_monthly_trend_sheet(writer)
                
                # 5. 데이터 입력 템플릿 시트
                self._create_input_template_sheet(writer)
            
            # 엑셀 파일 스타일링
            self._apply_excel_styling(filename)
            
            logger.info(f"엑셀 보고서 생성 완료: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"엑셀 보고서 생성 오류: {e}")
            return None
    
    def _create_raw_data_sheet(self, records: List[TradeRecord], writer):
        """원본 데이터 시트 생성"""
        data = []
        for record in records:
            data.append({
                'ID': record.id,
                '날짜': record.date,
                '수출국': record.country_origin,
                '수입국': record.country_destination,
                '수출업체': record.company_exporter or '',
                '수입업체': record.company_importer or '',
                '제품코드': record.product_code,
                '제품명': record.product_description,
                '수량': record.quantity,
                '단위': record.unit,
                '금액(USD)': record.value_usd,
                '거래유형': record.trade_type,
                '등록일시': record.created_at
            })
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='원본데이터', index=False)
    
    def _create_country_summary_sheet(self, records: List[TradeRecord], writer):
        """국가별 요약 시트 생성"""
        country_data = {}
        
        for record in records:
            country = record.country_origin
            if country not in country_data:
                country_data[country] = {
                    '수출국': country,
                    '거래건수': 0,
                    '총금액(USD)': 0,
                    '평균금액(USD)': 0,
                    '총수량': 0,
                    '주요제품': set()
                }
            
            country_data[country]['거래건수'] += 1
            country_data[country]['총금액(USD)'] += record.value_usd or 0
            country_data[country]['총수량'] += record.quantity or 0
            country_data[country]['주요제품'].add(record.product_description or '')
        
        # 평균 계산 및 주요제품 문자열 변환
        summary_data = []
        for country, data in country_data.items():
            data['평균금액(USD)'] = data['총금액(USD)'] / data['거래건수'] if data['거래건수'] > 0 else 0
            data['주요제품'] = ', '.join(list(data['주요제품'])[:3])  # 상위 3개만
            summary_data.append(data)
        
        df = pd.DataFrame(summary_data)
        df = df.sort_values('총금액(USD)', ascending=False)
        df.to_excel(writer, sheet_name='국가별요약', index=False)
    
    def _create_product_summary_sheet(self, records: List[TradeRecord], writer):
        """제품별 요약 시트 생성"""
        product_data = {}
        
        for record in records:
            product = record.product_description or '미분류'
            if product not in product_data:
                product_data[product] = {
                    '제품명': product,
                    '제품코드': record.product_code,
                    '거래건수': 0,
                    '총금액(USD)': 0,
                    '평균단가(USD/kg)': 0,
                    '총수량(kg)': 0,
                    '주요수출국': set()
                }
            
            product_data[product]['거래건수'] += 1
            product_data[product]['총금액(USD)'] += record.value_usd or 0
            product_data[product]['총수량(kg)'] += record.quantity or 0
            product_data[product]['주요수출국'].add(record.country_origin or '')
        
        # 평균 단가 계산
        summary_data = []
        for product, data in product_data.items():
            if data['총수량(kg)'] > 0:
                data['평균단가(USD/kg)'] = data['총금액(USD)'] / data['총수량(kg)']
            data['주요수출국'] = ', '.join(list(data['주요수출국'])[:3])
            summary_data.append(data)
        
        df = pd.DataFrame(summary_data)
        df = df.sort_values('총금액(USD)', ascending=False)
        df.to_excel(writer, sheet_name='제품별요약', index=False)
    
    def _create_monthly_trend_sheet(self, writer):
        """월별 트렌드 시트 생성"""
        try:
            # 최근 12개월 데이터 가져오기
            records = self.db.get_latest_records(365)
            
            monthly_data = {}
            for record in records:
                month_key = record.date.strftime('%Y-%m') if record.date else 'Unknown'
                if month_key not in monthly_data:
                    monthly_data[month_key] = {
                        '년월': month_key,
                        '거래건수': 0,
                        '총금액(USD)': 0,
                        '총수량(kg)': 0,
                        '평균단가(USD/kg)': 0
                    }
                
                monthly_data[month_key]['거래건수'] += 1
                monthly_data[month_key]['총금액(USD)'] += record.value_usd or 0
                monthly_data[month_key]['총수량(kg)'] += record.quantity or 0
            
            # 평균 단가 계산
            trend_data = []
            for month, data in monthly_data.items():
                if data['총수량(kg)'] > 0:
                    data['평균단가(USD/kg)'] = data['총금액(USD)'] / data['총수량(kg)']
                trend_data.append(data)
            
            df = pd.DataFrame(trend_data)
            df = df.sort_values('년월')
            df.to_excel(writer, sheet_name='월별트렌드', index=False)
            
        except Exception as e:
            logger.error(f"월별 트렌드 시트 생성 오류: {e}")
    
    def _create_input_template_sheet(self, writer):
        """데이터 입력 템플릿 시트 생성"""
        template_data = [
            {
                '날짜': '2025-01-10',
                '수출국': '호주',
                '수입국': '한국',
                '수출업체': '예: Australian Macadamia Co.',
                '수입업체': '예: Korea Nuts Import Ltd.',
                '제품코드': '080250',
                '제품명': '예: Raw Macadamia Nuts',
                '수량': 1000,
                '단위': 'kg',
                '금액(USD)': 15000,
                '거래유형': 'export',
                '비고': '신규 데이터 입력 시 이 행을 복사하여 사용'
            }
        ]
        
        df = pd.DataFrame(template_data)
        df.to_excel(writer, sheet_name='데이터입력템플릿', index=False)
    
    def _apply_excel_styling(self, filename: str):
        """엑셀 파일에 스타일 적용"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(filename)
            
            # 헤더 스타일 정의
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 헤더 행 스타일 적용
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 열 너비 자동 조정
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            
        except Exception as e:
            logger.error(f"엑셀 스타일링 오류: {e}")
    
    def create_comparison_report(self, days: int = 7) -> str:
        """비교 분석 엑셀 보고서 생성"""
        try:
            filename = f"reports/macadamia_comparison_{datetime.now().strftime('%Y%m%d')}.xlsx"
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # 기간별 비교 데이터
                current_records = self.db.get_latest_records(days)
                previous_records = self.db.get_records_by_date_range(
                    datetime.now() - timedelta(days=days*2),
                    datetime.now() - timedelta(days=days)
                )
                
                # 현재 기간 데이터
                self._create_period_summary(current_records, writer, f'최근{days}일')
                
                # 이전 기간 데이터
                self._create_period_summary(previous_records, writer, f'이전{days}일')
                
                # 비교 분석
                self._create_comparison_analysis(current_records, previous_records, writer)
            
            self._apply_excel_styling(filename)
            return filename
            
        except Exception as e:
            logger.error(f"비교 보고서 생성 오류: {e}")
            return None
    
    def _create_period_summary(self, records: List[TradeRecord], writer, sheet_name: str):
        """기간별 요약 시트 생성"""
        if not records:
            return
        
        summary_data = []
        country_stats = {}
        
        for record in records:
            country = record.country_origin
            if country not in country_stats:
                country_stats[country] = {
                    '수출국': country,
                    '거래건수': 0,
                    '총금액(USD)': 0,
                    '총수량(kg)': 0
                }
            
            country_stats[country]['거래건수'] += 1
            country_stats[country]['총금액(USD)'] += record.value_usd or 0
            country_stats[country]['총수량(kg)'] += record.quantity or 0
        
        for data in country_stats.values():
            summary_data.append(data)
        
        df = pd.DataFrame(summary_data)
        df = df.sort_values('총금액(USD)', ascending=False)
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    def _create_comparison_analysis(self, current_records: List[TradeRecord], previous_records: List[TradeRecord], writer):
        """비교 분석 시트 생성"""
        current_total = sum(record.value_usd or 0 for record in current_records)
        previous_total = sum(record.value_usd or 0 for record in previous_records)
        
        growth_rate = ((current_total - previous_total) / previous_total * 100) if previous_total > 0 else 0
        
        analysis_data = [
            {
                '항목': '총 거래액(USD)',
                '현재 기간': current_total,
                '이전 기간': previous_total,
                '증감액': current_total - previous_total,
                '증감률(%)': round(growth_rate, 2)
            },
            {
                '항목': '총 거래 건수',
                '현재 기간': len(current_records),
                '이전 기간': len(previous_records),
                '증감액': len(current_records) - len(previous_records),
                '증감률(%)': round(((len(current_records) - len(previous_records)) / len(previous_records) * 100), 2) if len(previous_records) > 0 else 0
            }
        ]
        
        df = pd.DataFrame(analysis_data)
        df.to_excel(writer, sheet_name='비교분석', index=False)
